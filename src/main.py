import pymysql
import openpyxl
from datetime import datetime


def main():
    connect = pymysql.connect(host='localhost', user='stest', password='1234', db='python', charset='utf8')

    cursor = connect.cursor()

    # 데이터 베이스 생성

    # sql = "CREATE DATABASE python"
    # cursor.execute(sql)

    # 테이블 생성

    # sql = '''CREATE TABLE purchase (
    # id int(10) NOT NULL AUTO_INCREMENT PRIMARY KEY,
    # order_date timestamp,
    # region varchar(10),
    # rep varchar(40),
    # item varchar(40),
    # units int(10),
    # unit_cost float,
    # east_regulation varchar(10),
    # total float
    # )'''

    sql = '''INSERT INTO purchase (order_date, region, rep, item, units, unit_cost, east_regulation, total)
    VALUES (%s, %s, %s, %s, %s, %s, %s, %s )'''

    file = openpyxl.load_workbook('SampleData.xlsx')

    sheet = file.active

    for r in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        order_date = r[0].value
        region = r[1].value
        rep = r[2].value
        item = r[3].value
        units = r[4].value
        unit_cost = r[5].value
        east_regulation = r[6].value
        total = r[7].value

        values = (order_date, region, rep, item, units, unit_cost, east_regulation, total)
        print(values)
        cursor.execute(sql, values)
        connect.commit()

    connect.close()


if __name__ == '__main__':
    main()
